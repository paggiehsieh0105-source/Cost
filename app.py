"""
階段4：Streamlit 前端介面

執行方式：
    cd src
    streamlit run app.py

四個分頁：
    1. 報價清單     - list_projects()，可複選多筆匯出成Excel
    2. 新增/編輯專案 - 六大成本區塊輸入介面，即時試算
    3. 包材主檔     - 對應 Excel「包材資料表」工作表
    4. 共用設定     - 對應 Excel「設定」工作表
"""

import streamlit as st
import pandas as pd

import database as db
from cost_model import (
    Settings, PackagingMaterial, CostProject, RawMaterial,
    PackagingLine, LaborLine, ShippingLine, CartonLine,
    calculate_total_cost,
)

DB_PATH = "cost_calculator.db"
db.init_db(DB_PATH)

st.set_page_config(page_title="代工成本試算系統", layout="wide")

PAGES = ["報價清單", "新增/編輯專案", "包材主檔", "共用設定"]
page = st.sidebar.radio("功能選單", PAGES)
st.sidebar.caption("化妝品OEM/ODM代工成本試算系統 — 階段4 Streamlit介面")


# ============================================================
# 共用小工具
# ============================================================

def money(x: float) -> str:
    return f"{x:,.2f}"


def _safe_float(val, default=0.0):
    """
    安全轉換成float，正確處理空白格子。

    重要：data_editor的空白格子在pandas裡會變成 NaN，但 NaN 在Python判斷式裡是「真值」
    （bool(float('nan')) == True），所以像 `row.get("單價") or 0` 這種寫法在欄位空白時
    並不會變成預設值0，而是維持NaN，後續存進SQLite的NOT NULL欄位就會出錯
    （sqlite3對NaN的特殊處理規則：插入NaN到NOT NULL欄位會被當成NULL，導致IntegrityError）。
    所以一律要用 pd.isna() 明確檢查，不能用 `or` 偷懶判斷數值類欄位。
    """
    if val is None or pd.isna(val):
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _safe_int(val, default=0):
    return int(_safe_float(val, default))


def _safe_str(val, default=""):
    """安全轉換成字串，空白格子(NaN/None)回傳預設值，避免NaN被當成字串存進資料庫。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    s = str(val).strip()
    return s if s else default


# ---- Excel 報價單匯出共用樣式與函式 ----
EXPORT_FONT_NAME = "微軟正黑體"
EXPORT_BORDER_COLOR = "595959"  # 深灰色
EXPORT_CM_TO_INCH = 1 / 2.54
EXPORT_MARGIN_CM = 2
EXPORT_NUM_COLS = 5


def _write_quote_block(ws, start_row, project_id, customer_name, product_name, monthly_qty,
                        result, tier_labels, tier_qtys, tier_margins, tier_prices):
    """把一筆報價單的內容寫進指定工作表的 start_row 開始的位置，回傳寫完後的下一個可用列號。"""
    from openpyxl.styles import Font, Alignment, Border, Side

    thin_gray_border = Border(
        left=Side(style="thin", color=EXPORT_BORDER_COLOR),
        right=Side(style="thin", color=EXPORT_BORDER_COLOR),
        top=Side(style="thin", color=EXPORT_BORDER_COLOR),
        bottom=Side(style="thin", color=EXPORT_BORDER_COLOR),
    )

    def _write_row(row_idx, values, bold=False, title=False, number_format=None):
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=EXPORT_FONT_NAME, bold=bold, size=14 if title else 11)
            cell.border = thin_gray_border
            cell.alignment = Alignment(horizontal="center" if title else "left", vertical="center")
            if number_format and isinstance(val, (int, float)):
                cell.number_format = number_format

    r = start_row
    _write_row(r, [f"代工成本試算報價單　({project_id})"] + [""] * (EXPORT_NUM_COLS - 1), bold=True, title=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=EXPORT_NUM_COLS)
    ws.row_dimensions[r].height = 28
    r += 2

    info_rows = [
        ("專案編號", project_id),
        ("客戶名稱", customer_name),
        ("產品名稱", product_name),
        ("預估月產量(件)", monthly_qty),
    ]
    for label, val in info_rows:
        _write_row(r, [label, val] + [""] * (EXPORT_NUM_COLS - 2), bold=True)
        r += 1
    r += 1

    _write_row(r, ["成本拆解"] + [""] * (EXPORT_NUM_COLS - 1), bold=True)
    r += 1
    cost_rows = [
        ("原物料成本", result.raw_material_total),
        ("直接人工成本", result.labor_total),
        ("製造費用", result.overhead),
        ("物流運費(分攤至單件)", result.shipping_cost_per_unit),
        ("單件總成本", result.unit_cost),
        ("系統建議報價", result.price_final),
        ("毛利金額", result.margin_amount),
        ("預估月營收", result.monthly_revenue),
    ]
    for label, val in cost_rows:
        _write_row(r, [label, val] + [""] * (EXPORT_NUM_COLS - 2), number_format="#,##0.00")
        r += 1
    r += 1

    _write_row(r, ["級距報價"] + [""] * (EXPORT_NUM_COLS - 1), bold=True)
    r += 1
    _write_row(r, ["級距", "數量(件)", "毛利率", "單件報價", "總金額"], bold=True)
    r += 1
    for label, qty, margin, price in zip(tier_labels, tier_qtys, tier_margins, tier_prices):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=int(qty))
        ws.cell(row=r, column=3, value=margin)
        ws.cell(row=r, column=4, value=price)
        ws.cell(row=r, column=5, value=price * qty)
        for col_idx in range(1, EXPORT_NUM_COLS + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = Font(name=EXPORT_FONT_NAME, size=11)
            cell.border = thin_gray_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        r += 1

    return r  # 下一筆報價可以從這一列開始接著寫


def _setup_quote_sheet_page(ws):
    """套用統一的版面設定：A4、置中、邊界2cm、欄寬。"""
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH, right=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH,
        top=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH, bottom=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH,
    )
    for col_idx in range(1, EXPORT_NUM_COLS + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def build_quote_workbook_for_project(project, settings, catalog):
    """組裝單一專案的報價單Excel工作簿（單一工作表），回傳 BytesIO。"""
    import io
    from openpyxl import Workbook
    from openpyxl.worksheet.page import PageMargins

    result = calculate_total_cost(project, settings, catalog)
    tier_labels = ["低量", "中量(主檔)", "高量"]
    tier_qtys = [max(project.monthly_quantity - 500, 0), project.monthly_quantity, project.monthly_quantity + 500]
    tier_margins = [result.tier_margin_low, result.tier_margin_mid, result.tier_margin_high]
    tier_prices = [result.tier_price_low, result.tier_price_mid, result.tier_price_high]

    wb = Workbook()
    ws = wb.active
    ws.title = "報價單"
    ws.print_options.verticalCentered = True
    _setup_quote_sheet_page(ws)
    _write_quote_block(
        ws, 1, project.project_id, project.customer_name, project.product_name,
        project.monthly_quantity, result, tier_labels, tier_qtys, tier_margins, tier_prices,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_quote_workbook_for_projects(project_ids, db_path):
    """組裝多個專案的報價單Excel工作簿——做成『並排比較表』：
    左邊第一欄是項目名稱，每個專案各佔一欄（橫向排列），方便逐項對照比較。
    回傳 BytesIO。"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    settings = db.load_settings(db_path)
    catalog = db.load_packaging_catalog(db_path)

    thin_gray_border = Border(
        left=Side(style="thin", color=EXPORT_BORDER_COLOR),
        right=Side(style="thin", color=EXPORT_BORDER_COLOR),
        top=Side(style="thin", color=EXPORT_BORDER_COLOR),
        bottom=Side(style="thin", color=EXPORT_BORDER_COLOR),
    )

    def _style_cell(cell, bold=False, title=False, center=False, number_format=None):
        cell.font = Font(name=EXPORT_FONT_NAME, bold=bold, size=14 if title else 11)
        cell.border = thin_gray_border
        cell.alignment = Alignment(
            horizontal="center" if (title or center) else "left", vertical="center"
        )
        if number_format:
            cell.number_format = number_format

    # ---- 先把每個專案的計算結果都算出來 ----
    project_results = []
    for pid in project_ids:
        project = db.load_project(pid, db_path)
        result = calculate_total_cost(project, settings, catalog)
        project_results.append((project, result))

    num_projects = len(project_results)
    num_cols = 1 + num_projects  # 第1欄是項目名稱，其餘每個專案各一欄

    wb = Workbook()
    ws = wb.active
    ws.title = "報價比較表"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if num_projects > 3 else "portrait"
    ws.print_options.horizontalCentered = True
    from openpyxl.worksheet.page import PageMargins
    ws.page_margins = PageMargins(
        left=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH, right=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH,
        top=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH, bottom=EXPORT_MARGIN_CM * EXPORT_CM_TO_INCH,
    )

    # ---- 標題列（合併） ----
    r = 1
    title_cell = ws.cell(row=r, column=1, value="代工成本試算　報價比較表")
    _style_cell(title_cell, bold=True, title=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    ws.row_dimensions[r].height = 28
    r += 2

    # ---- 每個專案各佔一欄的標題列：專案編號 / 客戶名稱 / 產品名稱 / 預估月產量 ----
    header_specs = [
        ("項目", None),
        ("專案編號", lambda p, res: p.project_id),
        ("客戶名稱", lambda p, res: p.customer_name),
        ("產品名稱", lambda p, res: p.product_name),
        ("預估月產量(件)", lambda p, res: p.monthly_quantity),
    ]
    for label, getter in header_specs:
        if getter is None:
            cell = ws.cell(row=r, column=1, value=label)
            _style_cell(cell, bold=True)
        else:
            ws.cell(row=r, column=1, value=label)
            _style_cell(ws.cell(row=r, column=1), bold=True)
            for col_offset, (project, result) in enumerate(project_results):
                cell = ws.cell(row=r, column=2 + col_offset, value=getter(project, result))
                _style_cell(cell, bold=True, center=True)
        r += 1
    r += 1

    # ---- 成本拆解：每一列一個項目，橫向比較各專案 ----
    section_cell = ws.cell(row=r, column=1, value="成本拆解")
    _style_cell(section_cell, bold=True)
    for col_offset in range(num_projects):
        _style_cell(ws.cell(row=r, column=2 + col_offset), bold=True)
    r += 1

    cost_specs = [
        ("原物料成本", lambda res: res.raw_material_total),
        ("直接人工成本", lambda res: res.labor_total),
        ("製造費用", lambda res: res.overhead),
        ("物流運費(分攤至單件)", lambda res: res.shipping_cost_per_unit),
        ("單件總成本", lambda res: res.unit_cost),
        ("系統建議報價", lambda res: res.price_final),
        ("毛利金額", lambda res: res.margin_amount),
        ("預估月營收", lambda res: res.monthly_revenue),
    ]
    for label, getter in cost_specs:
        ws.cell(row=r, column=1, value=label)
        _style_cell(ws.cell(row=r, column=1))
        for col_offset, (project, result) in enumerate(project_results):
            cell = ws.cell(row=r, column=2 + col_offset, value=getter(result))
            _style_cell(cell, center=True, number_format="#,##0.00")
        r += 1
    r += 1

    # ---- 級距報價：每個級距兩列（數量／單件報價），橫向比較各專案 ----
    section_cell = ws.cell(row=r, column=1, value="級距報價")
    _style_cell(section_cell, bold=True)
    for col_offset in range(num_projects):
        _style_cell(ws.cell(row=r, column=2 + col_offset), bold=True)
    r += 1

    tier_defs = [
        ("低量", lambda p: max(p.monthly_quantity - 500, 0), lambda res: res.tier_price_low),
        ("中量(主檔)", lambda p: p.monthly_quantity, lambda res: res.tier_price_mid),
        ("高量", lambda p: p.monthly_quantity + 500, lambda res: res.tier_price_high),
    ]
    for tier_label, qty_getter, price_getter in tier_defs:
        ws.cell(row=r, column=1, value=f"{tier_label} - 數量(件)")
        _style_cell(ws.cell(row=r, column=1))
        for col_offset, (project, result) in enumerate(project_results):
            cell = ws.cell(row=r, column=2 + col_offset, value=int(qty_getter(project)))
            _style_cell(cell, center=True, number_format="#,##0")
        r += 1

        ws.cell(row=r, column=1, value=f"{tier_label} - 單件報價")
        _style_cell(ws.cell(row=r, column=1))
        for col_offset, (project, result) in enumerate(project_results):
            cell = ws.cell(row=r, column=2 + col_offset, value=price_getter(result))
            _style_cell(cell, center=True, number_format="#,##0.00")
        r += 1

    # ---- 欄寬 ----
    ws.column_dimensions["A"].width = 22
    for col_offset in range(num_projects):
        ws.column_dimensions[get_column_letter(2 + col_offset)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# 分頁 1：報價清單
# ============================================================

if page == "報價清單":
    st.title("📋 報價清單")
    projects = db.list_projects(DB_PATH)
    if not projects:
        st.info("目前沒有任何報價，請到「新增/編輯專案」建立第一筆。")
    else:
        df = pd.DataFrame(projects)
        df = df.rename(columns={
            "project_id": "專案編號", "customer_name": "客戶名稱",
            "product_name": "產品名稱", "monthly_quantity": "預估月產量",
            "created_at": "建立時間",
        })
        df.insert(0, "選取", False)

        st.caption("勾選「選取」欄位可複選多筆報價，下方按鈕會把選中的報價各自匯出成一張工作表，合併在同一份Excel檔裡")
        edited_df = st.data_editor(
            df, use_container_width=True, hide_index=True, disabled=df.columns.drop("選取"),
            key="quote_list_editor",
        )

        selected_ids = edited_df.loc[edited_df["選取"], "專案編號"].tolist()
        st.write(f"已選取 **{len(selected_ids)}** 筆報價")

        if st.button(
            f"📥 匯出選中的 {len(selected_ids)} 筆報價單(Excel)",
            type="primary", disabled=not selected_ids,
        ):
            excel_buffer = build_quote_workbook_for_projects(selected_ids, DB_PATH)
            st.download_button(
                "💾 點此下載合併報價單.xlsx",
                data=excel_buffer,
                file_name="合併報價單.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        st.divider()
        del_id = st.selectbox("選擇要刪除的專案編號", [""] + [p["project_id"] for p in projects])
        if st.button("🗑️ 刪除此專案", type="secondary", disabled=not del_id):
            db.delete_project(del_id, DB_PATH)
            st.success(f"已刪除專案 {del_id}")
            st.rerun()


# ============================================================
# 分頁 2：新增/編輯專案
# ============================================================

elif page == "新增/編輯專案":
    st.title("🧮 新增 / 編輯 代工成本試算案")
    st.caption("💡 下面所有欄位都包在一個表單裡，輸入過程中不會整頁重新整理，按下最下方「試算並儲存」才會一次送出")

    projects = db.list_projects(DB_PATH)
    project_ids = [p["project_id"] for p in projects]
    mode = st.radio("模式", ["新增專案", "讀取既有專案編輯"], horizontal=True)

    loaded_project = None
    if mode == "讀取既有專案編輯" and project_ids:
        chosen = st.selectbox("選擇專案編號", project_ids)
        if chosen:
            loaded_project = db.load_project(chosen, DB_PATH)

    catalog = db.load_packaging_catalog(DB_PATH)
    settings_for_regions = db.load_settings(DB_PATH)
    region_names = list(settings_for_regions.region_rates.keys())
    carton_size_names = list(settings_for_regions.carton_rates.keys())

    # 重要：用 instance_id 當 widget key 的一部分，
    # 這樣切換「選擇專案編號」時，欄位才會正確換成新專案的資料，
    # 不會被「Streamlit記住上一個專案的內容」卡住而讀取失敗。
    instance_id = loaded_project.project_id if loaded_project else "__new__"

    # ---- 一、專案基本資訊 ----
    # 注意：這個區塊故意放在表單(st.form)外面，目的是讓下面的「包材明細」
    # 可以在你選擇包材名稱之後，立刻就把單價/規格帶出來顯示（表單裡的元件要等按下送出按鈕，
    # 整頁才會重新計算，沒辦法做到「選了名稱→馬上看到單價」這種即時互動）。
    st.subheader("一、專案基本資訊")
    c1, c2, c3 = st.columns(3)
    project_id = c1.text_input(
        "專案編號", value=loaded_project.project_id if loaded_project else "OEM-2026-001",
        key=f"project_id_{instance_id}",
    )
    customer_name = c2.text_input(
        "客戶名稱", value=loaded_project.customer_name if loaded_project else "",
        key=f"customer_name_{instance_id}",
    )
    product_name = c3.text_input(
        "產品名稱", value=loaded_project.product_name if loaded_project else "",
        key=f"product_name_{instance_id}",
    )
    monthly_qty = st.number_input(
        "預估月產量(件)", min_value=0,
        value=int(loaded_project.monthly_quantity) if loaded_project else 10000,
        key=f"monthly_qty_{instance_id}",
    )

    # ---- 二、原物料成本 ----
    st.subheader("二、原物料成本 (Raw Materials ＋ 包材)")
    st.caption("原料已簡化為：直接輸入每件成品的整組配方成本，訂單量會在最後的「成本加總與報價試算結果」才乘進去算總營收")
    rm = loaded_project.raw_material if loaded_project else RawMaterial(10.0)
    rm_price = st.number_input("整組配方成本(元) — 每件成品", value=float(rm.unit_price), key=f"rm_price_{instance_id}")

    pkg_state_key = f"pkg_lines_data_{instance_id}"
    if pkg_state_key not in st.session_state:
        st.session_state[pkg_state_key] = pd.DataFrame(
            [
                {
                    "包材名稱": catalog.id_to_name(l.material_id) or "",
                    "用量": l.quantity,
                    "損耗率": l.loss_rate,
                }
                for l in (loaded_project.packaging_lines if loaded_project else [])
            ],
            columns=["包材名稱", "用量", "損耗率"],
        )

    material_names = catalog.all_names()
    st.markdown(
        "**包材明細**（請從「包材名稱」下拉選單挑選，可直接打字搜尋；資料來源為「包材主檔」分頁）\n\n"
        "⚠️ **「用量」請填「一件成品的實際使用量」**（例如一個瓶身、一個瓶蓋通常填1），"
        "**不是訂單總數量**——訂單總共要採購多少，系統會自動用「用量 × 預估月產量」幫你算，不需要自己填。"
    )
    pkg_df = st.data_editor(
        st.session_state[pkg_state_key],
        num_rows="dynamic",
        use_container_width=True,
        key=f"editor_pkg_lines_{instance_id}",
        column_config={
            "包材名稱": st.column_config.SelectboxColumn(
                "包材名稱", options=material_names, required=False,
                help="從「包材主檔」挑選，可直接打字搜尋",
            ),
            "用量": st.column_config.NumberColumn(
                "用量(一件成品的實際使用量)", min_value=0.0, step=0.1, default=1.0,
                help="這裡填「做一件成品」要用幾個這個包材，通常是1。請不要填訂單總數量！",
            ),
            "損耗率": st.column_config.NumberColumn(
                "損耗率", min_value=0.0, max_value=1.0, step=0.01, format="percent", default=0.0,
                help="請輸入百分比，例如損耗3%就輸入 0.03（畫面上會自動顯示成 3%）",
            ),
        },
    )
    st.session_state[pkg_state_key] = pkg_df

    # ---- 即時顯示對應的單價/規格(若名稱重複會列出多筆供確認) ----
    preview_rows = []
    for _, row in pkg_df.iterrows():
        name = _safe_str(row.get("包材名稱"))
        if not name:
            continue
        matches = catalog.find_all_by_name(name)
        if not matches:
            preview_rows.append({
                "包材名稱": name, "編號": "⚠️ 找不到此名稱，請確認「包材主檔」",
                "規格": "", "單價": None, "用量": row.get("用量"), "損耗率": row.get("損耗率"),
            })
        else:
            m = matches[0]
            note = "" if len(matches) == 1 else f"（⚠️名稱重複共{len(matches)}筆，會採用編號{m.material_id}這一筆）"
            preview_rows.append({
                "包材名稱": name, "編號": m.material_id + note,
                "規格": m.spec, "單價": m.unit_price,
                "用量": row.get("用量"), "損耗率": row.get("損耗率"),
            })
    if preview_rows:
        st.caption("已選擇的包材（單價/規格自動帶出，僅供確認，實際計算仍以上方表格的數字為準）：")
        preview_df = pd.DataFrame(preview_rows)
        st.dataframe(
            preview_df.style.format({"用量": "{:.2f}", "損耗率": "{:.1%}", "單價": "{:.2f}"}),
            use_container_width=True, hide_index=True,
        )

    if not material_names:
        st.warning("包材主檔目前是空的，請先到「包材主檔」分頁新增。")

    with st.form(f"project_form_{instance_id}"):

        # ---- 三、直接人工成本 ----
        st.subheader("三、直接人工成本 (Direct Labor)")
        labor_default = pd.DataFrame(
            [
                {"製程名稱": l.process_name, "批量總件數": l.batch_quantity,
                 "批量總時間(分)": l.batch_time_minutes, "參與人數": l.headcount}
                for l in (loaded_project.labor_lines if loaded_project else [])
            ],
            columns=["製程名稱", "批量總件數", "批量總時間(分)", "參與人數"],
        )
        st.markdown("**人工製程明細（每件加工時間 = 批量總時間 ÷ 批量總件數，自動計算）**")
        labor_df = st.data_editor(
            labor_default, num_rows="dynamic", use_container_width=True, key=f"editor_labor_lines_{instance_id}",
        )

        # ---- 四、物流運費 ----
        st.subheader("四、物流運費 (Shipping)")
        ship_default = pd.DataFrame(
            [
                {"運送方式": l.method, "專車區域": l.region or "", "出貨數量": l.quantity}
                for l in (loaded_project.shipping_lines if loaded_project else [{"method": "物流", "quantity": 1}])
            ] if loaded_project else [{"運送方式": "物流", "專車區域": "", "出貨數量": 1}],
            columns=["運送方式", "專車區域", "出貨數量"],
        )
        st.markdown(
            "**物流運費明細**（運送方式請從下拉選單選擇；若選「專車」，請在「專車區域」欄位選擇縣市分區；"
            "以下「出貨數量」是用來算這一行自己的運費，不是分攤用的分母）"
        )
        ship_df = st.data_editor(
            ship_default,
            num_rows="dynamic",
            use_container_width=True,
            key=f"editor_ship_lines_{instance_id}",
            column_config={
                "運送方式": st.column_config.SelectboxColumn(
                    "運送方式", options=["專車", "物流", "棧板"], required=False,
                ),
                "專車區域": st.column_config.SelectboxColumn(
                    "專車區域", options=region_names, required=False,
                    help="只有「運送方式」選「專車」時才需要選這裡",
                ),
                "出貨數量": st.column_config.NumberColumn("出貨數量", min_value=0.0, step=1.0),
            },
        )

        # ---- 紙箱用量（依尺寸計費，併入物流運費） ----
        carton_default = pd.DataFrame(
            [
                {"紙箱尺寸": l.size, "用量(箱)": l.quantity}
                for l in (loaded_project.carton_lines if loaded_project else [])
            ],
            columns=["紙箱尺寸", "用量(箱)"],
        )
        if not carton_size_names:
            st.warning("尚未設定任何紙箱尺寸，請先到「共用設定」分頁新增。")
        st.markdown("**紙箱用量（依尺寸計費，費用會併入物流運費總額）**")
        carton_df = st.data_editor(
            carton_default,
            num_rows="dynamic",
            use_container_width=True,
            key=f"editor_carton_lines_{instance_id}",
            column_config={
                "紙箱尺寸": st.column_config.SelectboxColumn(
                    "紙箱尺寸", options=carton_size_names,
                    help="不同尺寸的紙箱費率請到「共用設定」分頁調整",
                ),
                "用量(箱)": st.column_config.NumberColumn("用量(箱)", min_value=0.0, step=1.0),
            },
        )

        st.caption(
            f"💡 物流運費／紙箱費用的分攤分母，固定使用上方填的「預估月產量」（目前是 {monthly_qty:,.0f} 件），"
            f"不需要另外設定；上面表格裡的「出貨數量」「用量(箱)」只是用來算這一列自己的費用。"
        )

        st.divider()
        submitted = st.form_submit_button("🧮 試算並儲存", type="primary", use_container_width=True)

    # ---- 表單送出後才執行：組裝成 CostProject、計算、儲存 ----
    if not submitted:
        st.info("請填寫上面的表單，按下「🧮 試算並儲存」按鈕後才會計算結果並存進資料庫。")
        st.stop()

    settings = db.load_settings(DB_PATH)

    def df_to_packaging(df):
        out = []
        for _, row in df.iterrows():
            name = _safe_str(row.get("包材名稱"))
            if not name:
                continue
            material = catalog.find_by_name(name)
            if material is None:
                st.warning(f"包材名稱「{name}」在主檔裡找不到，這一列會被忽略")
                continue
            out.append(PackagingLine(
                material_id=material.material_id,
                quantity=_safe_float(row.get("用量")),
                loss_rate=_safe_float(row.get("損耗率")),
            ))
        return out

    def df_to_labor(df):
        out = []
        for _, row in df.iterrows():
            process_name = _safe_str(row.get("製程名稱"))
            if not process_name:
                continue
            out.append(LaborLine(
                process_name=process_name,
                batch_quantity=_safe_float(row.get("批量總件數")),
                batch_time_minutes=_safe_float(row.get("批量總時間(分)")),
                headcount=_safe_int(row.get("參與人數"), default=1),
            ))
        return out

    def df_to_shipping(df):
        out = []
        for _, row in df.iterrows():
            method = _safe_str(row.get("運送方式"))
            if not method:
                continue
            region = _safe_str(row.get("專車區域")) or None
            out.append(ShippingLine(method=method, quantity=_safe_float(row.get("出貨數量")), region=region))
        return out

    def df_to_carton(df):
        out = []
        for _, row in df.iterrows():
            size = _safe_str(row.get("紙箱尺寸"))
            if not size:
                continue
            out.append(CartonLine(size=size, quantity=_safe_float(row.get("用量(箱)"))))
        return out

    try:
        project = CostProject(
            project_id=project_id,
            customer_name=customer_name,
            product_name=product_name,
            monthly_quantity=monthly_qty,
            raw_material=RawMaterial(rm_price),
            packaging_lines=df_to_packaging(pkg_df),
            labor_lines=df_to_labor(labor_df),
            shipping_lines=df_to_shipping(ship_df),
            carton_lines=df_to_carton(carton_df),
        )
        result = calculate_total_cost(project, settings, catalog)
        calc_error = None
    except Exception as e:
        result = None
        calc_error = str(e)

    st.subheader("六、成本加總與報價試算結果")
    if calc_error:
        st.error(f"計算發生錯誤：{calc_error}")
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("原物料成本", money(result.raw_material_total))
        c2.metric("直接人工成本", money(result.labor_total))
        c3.metric("製造費用", money(result.overhead))
        c4.metric(
            "物流運費(分攤至單件)", money(result.shipping_cost_per_unit),
            help=f"出貨總額 {money(result.shipping_total)} 元 ÷ 訂單量(預估月產量)，換算成每件分攤的運費",
        )
        st.caption(
            f"💡 物流運費總額為 {money(result.shipping_total)} 元（含紙箱費用 {money(result.carton_total)} 元，整批/整單），"
            f"已換算成每件分攤 {money(result.shipping_cost_per_unit)} 元，才併入下方的單件總成本"
        )

        c1, c2, c3 = st.columns(3)
        c1.metric("單件總成本(系統自動算)", money(result.unit_cost))
        c2.metric("系統建議報價", money(result.price_final))
        c3.metric("毛利金額(依系統建議報價)", money(result.margin_amount))

        st.metric("預估月營收(依系統建議報價 × 預估月產量)", money(result.monthly_revenue))

        # ---- 級距報價結果（低量／高量毛利率直接輸入） ----
        st.subheader("五、級距報價（低量／高量毛利率為你在「共用設定」頁直接輸入的數字）")
        st.caption(
            f"毛利率計算模式：{settings.margin_mode}；主檔毛利率 {settings.margin_rate:.0%}，"
            f"低量毛利率 {settings.tier_margin_low:.0%}、高量毛利率 {settings.tier_margin_high:.0%}（可在「共用設定」頁調整）"
        )
        tier_labels = ["低量", "中量(主檔)", "高量"]
        tier_qtys = [max(monthly_qty - 500, 0), monthly_qty, monthly_qty + 500]
        tier_margins = [result.tier_margin_low, result.tier_margin_mid, result.tier_margin_high]
        tier_prices = [result.tier_price_low, result.tier_price_mid, result.tier_price_high]
        tier_rows = []
        for label, qty, margin, price in zip(tier_labels, tier_qtys, tier_margins, tier_prices):
            tier_rows.append({
                "級距": label,
                "數量(件)": int(qty),
                "毛利率": margin,
                "單件報價": price,
                "總金額": price * qty,
            })
        tier_df = pd.DataFrame(tier_rows)
        st.dataframe(
            tier_df.style.format({"毛利率": "{:.1%}", "單件報價": "{:,.2f}", "總金額": "{:,.2f}"}),
            use_container_width=True, hide_index=True,
        )

        # ---- 列印 / 匯出 ----
        st.subheader("列印或匯出")
        st.caption("匯出Excel報表，內含專案資訊、成本拆解、級距報價；若要列印，可直接用瀏覽器的列印功能（Ctrl+P / Cmd+P）")

        excel_buffer = build_quote_workbook_for_project(project, settings, catalog)

        st.download_button(
            "📥 下載Excel報價單",
            data=excel_buffer,
            file_name=f"{project_id}_報價單.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        db.save_project(project, DB_PATH)
        st.success(f"專案 {project_id} 已儲存！")


# ============================================================
# 分頁 3：包材主檔
# ============================================================

elif page == "包材主檔":
    st.title("📦 包材主檔")
    st.caption("此頁對應 Excel「包材資料表」工作表，是全廠共用的包材清單。輸入過程不會整頁重整，按下按鈕才送出。")

    catalog = db.load_packaging_catalog(DB_PATH)
    existing = pd.DataFrame(
        [
            {"包材編號": m.material_id, "包材名稱": m.name, "規格": m.spec,
             "單位": m.unit, "單價": m.unit_price, "供應商": m.vendor, "備註": m.note}
            for m in catalog.all_materials()
        ],
        columns=["包材編號", "包材名稱", "規格", "單位", "單價", "供應商", "備註"],
    )

    with st.form("packaging_master_form"):
        st.markdown(
            "**包材清單**（可直接增刪修改，按下方按鈕儲存；「單價」請務必填數字，不可留空；"
            "「包材名稱」可以重複（例如不同供應商同一品名），但「包材編號」務必唯一，建議自己填，不要留空）"
        )
        df = st.data_editor(
            existing, num_rows="dynamic", use_container_width=True, key="editor_pkg_master_df",
            column_config={
                "單價": st.column_config.NumberColumn("單價", min_value=0.0, step=0.1, format="%.2f"),
            },
        )
        save_clicked = st.form_submit_button("💾 儲存包材主檔變更", type="primary")

    if save_clicked:
        skipped_rows = []
        used_ids = set(catalog.all_ids())
        for idx, row in df.iterrows():
            name = _safe_str(row.get("包材名稱"))
            if not name:
                continue
            if pd.isna(row.get("單價")):
                skipped_rows.append(name)
                continue
            mid = _safe_str(row.get("包材編號"))
            if not mid:
                # 沒填編號：自動產生一個，並確保不會跟既有/這批新存的其他筆撞號
                base = f"PKG-{name}"
                mid = base
                suffix = 1
                while mid in used_ids:
                    suffix += 1
                    mid = f"{base}-{suffix}"
            used_ids.add(mid)
            db.upsert_packaging_material(
                PackagingMaterial(
                    material_id=mid, name=name, spec=_safe_str(row.get("規格")),
                    unit=_safe_str(row.get("單位")), unit_price=_safe_float(row.get("單價")),
                    vendor=_safe_str(row.get("供應商")), note=_safe_str(row.get("備註")),
                ),
                DB_PATH,
            )
        if skipped_rows:
            st.warning(f"以下包材因為「單價」是空白，沒有儲存：{', '.join(skipped_rows)}，請補上單價後再存一次")
        st.success("包材主檔已更新！")
        st.rerun()


# ============================================================
# 分頁 4：共用設定
# ============================================================

elif page == "共用設定":
    st.title("⚙️ 共用設定")
    st.caption("此頁對應 Excel「設定」工作表，全廠只有一份共用參數。輸入過程不會整頁重整，按下按鈕才送出。")

    settings = db.load_settings(DB_PATH)

    with st.form("settings_form"):
        st.subheader("毛利與稅金")
        c1, c2 = st.columns(2)
        margin_mode = c1.selectbox(
            "毛利率計算模式", ["成本基礎(加成法)", "售價基礎(毛利率法)"],
            index=0 if settings.margin_mode == "成本基礎(加成法)" else 1,
        )
        margin_rate = c2.number_input("毛利率", value=float(settings.margin_rate), step=0.01, format="%.2f")
        c1, c2 = st.columns(2)
        overhead_rate = c1.number_input("製造費用攤提率", value=float(settings.overhead_rate), step=0.01, format="%.2f")
        tax_rate = c2.number_input("營業稅率", value=float(settings.tax_rate), step=0.01, format="%.2f")
        tax_mode = st.selectbox("報價是否含稅", ["未稅", "含稅"], index=0 if settings.tax_mode == "未稅" else 1)

        st.subheader("損耗率設定（％）")
        st.caption(
            "分別套用在原物料／人工／物流，算完各自的成本小計之後，再乘以(1+損耗率)當整體緩衝，"
            "代表「實際因為製程損耗，大概還要多抓多少%才夠」。請用小數表示，例如5%請輸入0.05。"
        )
        c1, c2, c3 = st.columns(3)
        raw_material_loss_rate = c1.number_input(
            "原物料損耗率", value=float(settings.raw_material_loss_rate),
            min_value=0.0, max_value=1.0, step=0.01, format="%.2f",
        )
        labor_loss_rate = c2.number_input(
            "人工損耗率", value=float(settings.labor_loss_rate),
            min_value=0.0, max_value=1.0, step=0.01, format="%.2f",
        )
        shipping_loss_rate = c3.number_input(
            "物流損耗率", value=float(settings.shipping_loss_rate),
            min_value=0.0, max_value=1.0, step=0.01, format="%.2f",
        )

        st.subheader("級距報價設定")
        st.caption("低量／高量級距請直接輸入你想要的毛利率，中量(主檔)沿用上面的「毛利率」")
        c1, c2 = st.columns(2)
        tier_margin_low = c1.number_input(
            "低量級距毛利率", value=float(settings.tier_margin_low),
            min_value=0.0, max_value=0.95, step=0.01, format="%.2f",
        )
        tier_margin_high = c2.number_input(
            "高量級距毛利率", value=float(settings.tier_margin_high),
            min_value=0.0, max_value=0.95, step=0.01, format="%.2f",
        )
        st.caption(
            f"套用公式會依照上面選的「毛利率計算模式」（{margin_mode}）："
            f"低量報價 = 用毛利率{tier_margin_low:.0%}算；中量(主檔)報價 = 用毛利率{margin_rate:.0%}算；"
            f"高量報價 = 用毛利率{tier_margin_high:.0%}算"
        )

        st.subheader("人工時薪換算")
        c1, c2, c3 = st.columns(3)
        monthly_salary = c1.number_input("作業員月薪", value=float(settings.monthly_salary))
        work_days = c2.number_input("每月工作天數", value=int(settings.work_days_per_month))
        work_hours = c3.number_input("每日工作時數", value=float(settings.work_hours_per_day))
        preview_wage = monthly_salary / (work_days * work_hours) if work_days and work_hours else 0
        st.caption(f"換算時薪 = {money(preview_wage)} 元/小時")

        st.subheader("物流運費")
        c1, c2 = st.columns(2)
        logistics_rate = c1.number_input("物流每件運費", value=float(settings.logistics_rate))
        pallet_rate = c2.number_input("棧板每件運費", value=float(settings.pallet_rate))

        st.markdown("**專車－依縣市區域報價**")
        region_rate_inputs = {}
        cols = st.columns(len(settings.region_rates))
        for col, (region, rate) in zip(cols, settings.region_rates.items()):
            region_rate_inputs[region] = col.number_input(region, value=float(rate))

        st.subheader("紙箱費率（依尺寸計費）")
        carton_df = pd.DataFrame(
            [{"紙箱尺寸": size, "每箱費用(元)": rate} for size, rate in settings.carton_rates.items()],
            columns=["紙箱尺寸", "每箱費用(元)"],
        )
        carton_rate_df = st.data_editor(
            carton_df, num_rows="dynamic", use_container_width=True, key="editor_carton_rates",
            column_config={
                "每箱費用(元)": st.column_config.NumberColumn("每箱費用(元)", min_value=0.0, step=1.0, format="%.2f"),
            },
        )
        st.caption("可直接新增/修改/刪除紙箱尺寸，「新增/編輯專案」頁的紙箱用量下拉選單會自動同步")

        save_clicked = st.form_submit_button("💾 儲存設定", type="primary", use_container_width=True)

    if save_clicked:
        carton_rates = {}
        for _, row in carton_rate_df.iterrows():
            size = _safe_str(row.get("紙箱尺寸"))
            if not size:
                continue
            carton_rates[size] = _safe_float(row.get("每箱費用(元)"))

        new_settings = Settings(
            margin_mode=margin_mode, margin_rate=margin_rate, overhead_rate=overhead_rate,
            tax_rate=tax_rate, tax_mode=tax_mode, monthly_salary=monthly_salary,
            work_days_per_month=int(work_days), work_hours_per_day=work_hours,
            logistics_rate=logistics_rate, pallet_rate=pallet_rate, region_rates=region_rate_inputs,
            carton_rates=carton_rates, tier_margin_low=tier_margin_low, tier_margin_high=tier_margin_high,
            raw_material_loss_rate=raw_material_loss_rate, labor_loss_rate=labor_loss_rate,
            shipping_loss_rate=shipping_loss_rate,
        )
        db.save_settings(new_settings, DB_PATH)
        st.success("設定已儲存！")
